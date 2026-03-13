"""
Zentrade — Generador de Calendario de Contenido Pre-Lanzamiento
4 semanas, 3 plataformas: X/Twitter, Instagram, TikTok
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import date

# ---------------------------------------------------------------------------
# DATOS DE CONTENIDO — 4 semanas × plataformas
# Asumimos que la Semana 1 arranca el lunes 2026-03-09
# ---------------------------------------------------------------------------

content = [

    # =========================================================
    # SEMANA 1 — PRESENCIA (NO VENTA)
    # =========================================================

    # ---- X / Twitter ----
    {
        "semana": "Semana 1 — Presencia",
        "fecha": "2026-03-09",
        "plataforma": "X / Twitter",
        "pilar": "Educativo",
        "formato": "Tweet standalone",
        "hook_titulo": "La razón #1 por la que los traders fallan evaluaciones (no es lo que crees)",
        "copy": (
            "La razón #1 por la que los traders fallan evaluaciones de prop firms no es la estrategia.\n\n"
            "Es el revenge trading.\n\n"
            "Pierdes un trade → entras al siguiente sin plan → vuelas tu cuenta en 20 minutos.\n\n"
            "El drawdown no te mata. Tu reacción al drawdown sí.\n\n"
            "¿Cuántas evaluaciones has fallado así? Comenta abajo."
        ),
        "direccion_visual": (
            "Fondo oscuro #0D0D0D. Texto blanco limpio. Sin gráficas — solo texto con espaciado. "
            "Opcional: un separador de línea verde neón (#00FF87) entre párrafos si se hace imagen."
        ),
        "hashtags": "#propfirm #FTMO #TopStep #tradingpsychology #futurestrading",
        "publicacion": "Lunes 09:00 AM ET",
        "hook_alternativo": "Fallé 4 evaluaciones antes de entender esto sobre el revenge trading →",
        "notas_marca": "Tono: confidente, sin promocionar Zentrade. Terminar con pregunta para generar replies.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 1 — Presencia",
        "fecha": "2026-03-10",
        "plataforma": "X / Twitter",
        "pilar": "Behind the scenes",
        "formato": "Tweet standalone",
        "hook_titulo": "Por qué construí un journal de trading siendo developer",
        "copy": (
            "Llevo 2 años operando futuros.\n\n"
            "Probé Excel, Notion, TradeZella — nada funcionaba como yo quería.\n\n"
            "Excel: demasiado manual.\n"
            "TradeZella: caro, sin foco en prop firms, sin español.\n"
            "Notion: bonito pero inútil para métricas reales.\n\n"
            "Así que construí lo que necesitaba.\n\n"
            "Se llama @zentrade_app — y en unas semanas lo van a poder usar."
        ),
        "direccion_visual": (
            "Tweet de texto puro. Sin imagen necesaria. Si se añade imagen, usar screenshot difuminado "
            "del dashboard de Zentrade con blur leve para generar intriga."
        ),
        "hashtags": "#buildinpublic #propfirm #tradingjournal #futurestrading",
        "publicacion": "Martes 10:00 AM ET",
        "hook_alternativo": "Gasté 2 años buscando el journal perfecto para prop firms. No existía. Lo construí.",
        "notas_marca": "Voz del fundador. Auténtico, no corporativo. Establece credibilidad desde el día 1.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 1 — Presencia",
        "fecha": "2026-03-12",
        "plataforma": "X / Twitter",
        "pilar": "Educativo",
        "formato": "Thread (5 tweets)",
        "hook_titulo": "5 métricas que toda prop firm revisa — y que la mayoría de traders ignora [hilo]",
        "copy": (
            "--- TWEET 1 (HOOK) ---\n"
            "5 métricas que toda prop firm revisa — y que la mayoría de traders ignora.\n\n"
            "Si no las estás midiendo, estás operando a ciegas.\n\n"
            "Hilo:\n\n"
            "--- TWEET 2 ---\n"
            "1. Consistency %\n\n"
            "No basta con ser rentable. Debes serlo de forma CONSISTENTE.\n\n"
            "La regla: ningún día debe generar más del 30-50% de tu ganancia total del periodo.\n\n"
            "Un día de suerte no es una cuenta fondeada. Es una trampa.\n\n"
            "--- TWEET 3 ---\n"
            "2. Profit Factor\n\n"
            "Fórmula: Ganancia bruta ÷ Pérdida bruta\n\n"
            "Menos de 1.0 = estrategia perdedora.\n"
            "Entre 1.0 y 1.5 = sobreviviendo.\n"
            "Más de 1.5 = vas por buen camino.\n\n"
            "La mayoría no sabe cuál es el suyo hasta que falla la evaluación.\n\n"
            "--- TWEET 4 ---\n"
            "3. Max Daily Loss usado\n\n"
            "¿Qué tan cerca del límite estás operando cada día?\n\n"
            "Si regularmente llegas al 80-90% de tu MDL antes de parar = señal de alerta.\n\n"
            "Mídelo. No lo adivines.\n\n"
            "--- TWEET 5 ---\n"
            "4. Win Rate vs. Risk/Reward\n\n"
            "Un win rate del 40% puede ser rentable con RR 1:2.5\n"
            "Un win rate del 70% puede hacerte perder dinero con RR 1:0.5\n\n"
            "Necesitas los dos números juntos para que signifiquen algo.\n\n"
            "--- TWEET 6 ---\n"
            "5. Drawdown relativo vs. trailing drawdown\n\n"
            "Son diferentes. Confundirlos es el error más caro en evaluaciones tipo Apex o TopStep.\n\n"
            "¿Los estás midiendo por separado?\n\n"
            "Si quieres un dashboard que calcule todo esto automáticamente → sigue @zentrade_app.\n\n"
            "Lanzamos pronto."
        ),
        "direccion_visual": (
            "Thread solo texto. En el último tweet, si se añade imagen, usar screenshot del KPI dashboard "
            "de Zentrade mostrando profit factor y consistency % en dark mode."
        ),
        "hashtags": "#propfirm #FTMO #ApexTrader #TopStep #tradingjournal #futurestrading #fundedtrader",
        "publicacion": "Jueves 09:00 AM ET",
        "hook_alternativo": "¿Sabes cuál es tu Profit Factor? Si la respuesta es 'no estoy seguro', lee esto →",
        "notas_marca": "El thread educativo de más valor de la semana. Pinear si tiene buen engagement. Terminar con CTA suave (seguir la cuenta).",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 1 — Presencia",
        "fecha": "2026-03-13",
        "plataforma": "X / Twitter",
        "pilar": "Comunidad",
        "formato": "Tweet pregunta",
        "hook_titulo": "¿Cuántas evaluaciones has fallado antes de pasar tu primera?",
        "copy": (
            "¿Cuántas evaluaciones de prop firm has fallado antes de pasar la primera?\n\n"
            "A) 0-2\n"
            "B) 3-5\n"
            "C) 6-10\n"
            "D) Más de 10 (sin juicio, estamos aquí)\n\n"
            "RT si crees que el problema no era la estrategia."
        ),
        "direccion_visual": (
            "Tweet de texto puro. Sin imagen. Diseñado para maximizar replies y retweets."
        ),
        "hashtags": "#propfirm #FTMO #fundedtrader #futurestrading",
        "publicacion": "Viernes 02:00 PM ET",
        "hook_alternativo": "Seamos honestos: ¿cuánto has gastado en evaluaciones de prop firms que fallaste?",
        "notas_marca": "Engagement post. Responder TODOS los comentarios el mismo día para construir comunidad desde cero.",
        "estado": "Borrador",
    },

    # ---- Instagram ----
    {
        "semana": "Semana 1 — Presencia",
        "fecha": "2026-03-10",
        "plataforma": "Instagram",
        "pilar": "Educativo",
        "formato": "Carrusel (5 slides)",
        "hook_titulo": "5 errores que te hacen fallar la evaluación (aunque tu estrategia sea buena)",
        "copy": (
            "--- SLIDE 1 (PORTADA) ---\n"
            "5 errores que te hacen fallar la evaluación\n"
            "(aunque tu estrategia sea buena)\n\n"
            "--- SLIDE 2 ---\n"
            "Error #1: Operar sin journal\n\n"
            "No sabes qué funciona, qué no, ni por qué.\n"
            "Estás operando con feeling, no con datos.\n\n"
            "--- SLIDE 3 ---\n"
            "Error #2: Ignorar la regla de consistencia\n\n"
            "Un día de $800 en una cuenta de $50K puede invalidar todo tu avance.\n"
            "Las prop firms no solo quieren que seas rentable. Quieren que seas predecible.\n\n"
            "--- SLIDE 4 ---\n"
            "Error #3: Revenge trading después de una pérdida\n\n"
            "Pierdes 1 trade → entras sin setup → pierdes 3 más.\n"
            "20 minutos pueden destruir semanas de trabajo.\n\n"
            "--- SLIDE 5 ---\n"
            "Error #4: Sobretrading en días malos\n\n"
            "Más trades ≠ más ganancias.\n"
            "Los mejores traders funded tienen 2-3 trades al día. Con criterio.\n\n"
            "--- SLIDE 6 (CTA) ---\n"
            "Error #5: No revisar tus métricas clave antes de operar\n\n"
            "Profit factor, win rate, drawdown actual — si no los revisas a diario, estás adivinando.\n\n"
            "Guarda este carrusel para tu próxima sesión de trading.\n"
            "Seguinos @zentrade_app — estamos construyendo la herramienta para evitar exactamente esto."
        ),
        "direccion_visual": (
            "Fondo oscuro #0D0D0D en todas las slides. Tipografía sans-serif blanca (Inter o similar). "
            "Número del error en verde neón (#00FF87) grande. "
            "Slide 1: logo @zentrade_app pequeño en esquina inferior. "
            "Línea de acento verde horizontal en cada slide. "
            "Formato cuadrado 1:1 para feed. Línea final con 'Desliza →' en slide 1."
        ),
        "hashtags": "#propfirm #FTMO #TopStep #ApexTrader #tradingjournal #futurestrading #fundedtrader #tradingpsychology #NinjaTrader #Tradovate",
        "publicacion": "Martes 07:00 PM hora Colombia/México",
        "hook_alternativo": "Tu estrategia no es el problema. Estos 5 hábitos sí lo son →",
        "notas_marca": "Carrusel de mayor valor de la semana. Diseñado para guardar/compartir. Terminar con CTA de seguir la cuenta.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 1 — Presencia",
        "fecha": "2026-03-12",
        "plataforma": "Instagram",
        "pilar": "Producto en acción",
        "formato": "Reel (30-45 seg)",
        "hook_titulo": "Así se ve un dashboard de trading hecho para prop firms",
        "copy": (
            "--- GUIÓN DEL REEL ---\n\n"
            "[00:00-00:03] HOOK VISUAL: Pantalla oscura, dashboard aparece con animación. "
            "Texto superpuesto: 'Esto es lo que los traders funded revisan cada día.'\n\n"
            "[00:03-00:15] Narración en voz: 'Si estás en evaluación, hay 3 números que no puedes ignorar: "
            "tu consistency score, tu profit factor, y cuánto drawdown llevas. "
            "Aquí los tienes en tiempo real, en un solo lugar.'\n\n"
            "[00:15-00:30] Screen recording: hover sobre KPIs, mostrar equity curve, mostrar calendario "
            "con días verdes/rojos. Sin audio marketero — música lo-fi tranquila de fondo.\n\n"
            "[00:30-00:40] Texto final: '@zentrade_app — lanzamos pronto. "
            "Link en bio para ser de los primeros.'\n\n"
            "--- CAPTION ---\n\n"
            "Así debería verse tu journal de trading si vas en serio con las prop firms.\n\n"
            "No Excel. No Notion. Un dashboard construido para lo que las prop firms realmente miden.\n\n"
            "Lanzamos pronto. Link en bio para acceso anticipado.\n"
            "@zentrade_app"
        ),
        "direccion_visual": (
            "Screen recording del dashboard en dark mode. Resolución nítida. "
            "Música lo-fi instrumental, sin voz si no hay locutor disponible — usar texto superpuesto. "
            "Color de texto overlay: blanco o verde neón. "
            "Primer frame: debe tener hook visual impactante — dashboard con métricas reales (usar datos demo). "
            "Formato vertical 9:16 para Reels."
        ),
        "hashtags": "#propfirm #tradingjournal #fundedtrader #futurestrading #FTMO #TopStep #tradingapp",
        "publicacion": "Jueves 06:00 PM hora Colombia/México",
        "hook_alternativo": "3 métricas que separan a los traders funded de los que siguen fallando evaluaciones →",
        "notas_marca": "Primer reel de producto. No vender — mostrar. El dashboard debe verse moderno y limpio. Usar datos demo reales para credibilidad.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 1 — Presencia",
        "fecha": "2026-03-14",
        "plataforma": "Instagram",
        "pilar": "Behind the scenes",
        "formato": "Post imagen + caption",
        "hook_titulo": "Por qué construimos Zentrade (historia del fundador)",
        "copy": (
            "Esto no nació de un PowerPoint ni de una idea de negocio.\n\n"
            "Nació de perder evaluaciones que debería haber pasado.\n\n"
            "Tenía la estrategia. Tenía el riesgo controlado. Pero no tenía claridad sobre mis propios patrones: "
            "cuándo operaba bien, cuándo me saboteaba, qué días del mes eran mis peores días.\n\n"
            "Probé todo lo que existía. Nada estaba hecho para traders de futuros en evaluación.\n\n"
            "Así que lo construimos nosotros.\n\n"
            "Zentrade. Un journal con IA diseñado para traders que van en serio con las prop firms.\n\n"
            "Seguinos. Lanzamos pronto.\n\n"
            "@zentrade_app"
        ),
        "direccion_visual": (
            "Imagen: foto del fundador en setup de trading (fondo oscuro, pantallas de trading visibles) "
            "O imagen artística de un setup de trading dark mode con texto superpuesto. "
            "Si no hay foto disponible: fondo degradado oscuro con tipografía de la frase clave: "
            "'No tenía claridad sobre mis propios patrones. Así que lo construimos nosotros.' "
            "Formato cuadrado 1:1."
        ),
        "hashtags": "#buildinpublic #propfirm #futurestrading #tradingjournal #fundedtrader #tradingpsychology",
        "publicacion": "Sábado 11:00 AM hora Colombia/México",
        "hook_alternativo": "Fallé evaluaciones que debería haber pasado. Este es el porqué de Zentrade.",
        "notas_marca": "Post de conexión emocional. Establece la narrativa del fundador desde el día 1. Auténtico > pulido.",
        "estado": "Borrador",
    },

    # ---- TikTok ----
    {
        "semana": "Semana 1 — Presencia",
        "fecha": "2026-03-11",
        "plataforma": "TikTok",
        "pilar": "Educativo",
        "formato": "Video (45-60 seg)",
        "hook_titulo": "La regla de consistencia que nadie te explica en las prop firms",
        "copy": (
            "--- GUIÓN TIKTOK ---\n\n"
            "[00:00-00:03] HOOK HABLADO (a cámara o voz en off):\n"
            "'Si estás en evaluación de prop firm, hay una regla que puede hacerte fallar incluso si eres rentable.'\n\n"
            "[00:03-00:20] DESARROLLO:\n"
            "'Se llama la regla de consistencia. La mayoría de prop firms — FTMO, Apex, TopStep — "
            "exigen que ningún día genere más del 30 o 50 por ciento de tu ganancia total del periodo. "
            "O sea: si ganaste 2,000 dólares en la evaluación y 1,200 fueron en un solo día... rechazado. "
            "Aunque hayas pasado el profit target.'\n\n"
            "[00:20-00:40] DATO/EJEMPLO:\n"
            "'Imagina esto: llevas 3 semanas operando. Profit target alcanzado. "
            "Pero ese día que tuviste una racha increíble y sacaste 800 dólares en 2 horas — "
            "ese día representa el 45% de tu ganancia total. Reprobado.'\n\n"
            "[00:40-00:55] CTA:\n"
            "'¿Estás midiendo tu consistency score? Comenta cuál es tu prop firm y te explico "
            "cómo aplica la regla para tu cuenta. Sígueme para más tips de evaluación.'\n\n"
            "--- TEXTO SUPERPUESTO EN PANTALLA ---\n"
            "0:00 → 'Puedes FALLAR aunque seas rentable'\n"
            "0:10 → 'Regla de consistencia = ningún día > 30-50% del profit total'\n"
            "0:35 → 'Ejemplo: $800 de $1,800 = 44% → RECHAZADO'"
        ),
        "direccion_visual": (
            "Formato vertical 9:16. Fondo: pantalla de trading o setup oscuro. "
            "Texto superpuesto en blanco/verde sobre fondo semitransparente. "
            "Alternativa: talking head a cámara con overlay de texto. "
            "Música de fondo: trending en TikTok trader niche (lo-fi o beat tranquilo). "
            "Sin efectos excesivos — credibilidad sobre entretenimiento."
        ),
        "hashtags": "#propfirm #FTMO #TopStep #ApexTrader #futurestrading #fundedtrader #tradingpsychology #tradertips",
        "publicacion": "Miércoles 07:00 PM hora Colombia/México",
        "hook_alternativo": "Pasé el profit target y me rechazaron igual. Esta fue la razón.",
        "notas_marca": "TikTok educativo. Comentar en respuestas para alimentar el algoritmo. Responder con video si hay dudas frecuentes.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 1 — Presencia",
        "fecha": "2026-03-14",
        "plataforma": "TikTok",
        "pilar": "Behind the scenes",
        "formato": "Video (30-45 seg)",
        "hook_titulo": "Construyendo un SaaS de trading desde cero — día 1 público",
        "copy": (
            "--- GUIÓN TIKTOK ---\n\n"
            "[00:00-00:03] HOOK:\n"
            "'Este es el día que empiezo a mostrar cómo construí una app de trading en público.'\n\n"
            "[00:03-00:25] STORY:\n"
            "'Soy developer y trader de futuros. Durante 2 años busqué un journal que funcionara "
            "para traders en evaluación de prop firms. Excel era un caos. Las opciones que existían "
            "eran caras, en inglés, y no estaban pensadas para lo que las prop firms realmente miden. "
            "Así que lo construí yo mismo.'\n\n"
            "[00:25-00:40] REVEAL + CTA:\n"
            "'Se llama Zentrade. Lanzamos en semanas. "
            "Sígueme si quieres ver cómo va el build — y cómo puedes ser de los primeros en usarlo.'\n\n"
            "--- TEXTO SUPERPUESTO ---\n"
            "0:00 → 'Construyendo un journal de trading con IA'\n"
            "0:15 → 'Porque lo que existía no era suficiente'\n"
            "0:30 → '@zentrade_app — lanzamos pronto'"
        ),
        "direccion_visual": (
            "Selfie a cámara con setup de trading de fondo, o screen recording del dashboard con voz en off. "
            "Fondo: pantallas de trading, dark mode. "
            "Muy auténtico — no producción alta. El buildinpublic content funciona mejor sin edición exagerada."
        ),
        "hashtags": "#buildinpublic #saas #propfirm #futurestrading #tradingapp #indiedev #fundedtrader",
        "publicacion": "Sábado 12:00 PM hora Colombia/México",
        "hook_alternativo": "Tardé 2 años en encontrar el journal perfecto para prop firms. Al final lo tuve que construir.",
        "notas_marca": "Build in public genera audiencia de traders Y de developers interesados en SaaS. Doble audiencia.",
        "estado": "Borrador",
    },

    # =========================================================
    # SEMANA 2 — CONSTRUIR AUDIENCIA ACTIVA
    # =========================================================

    # ---- X / Twitter ----
    {
        "semana": "Semana 2 — Audiencia activa",
        "fecha": "2026-03-16",
        "plataforma": "X / Twitter",
        "pilar": "Educativo",
        "formato": "Thread (6 tweets)",
        "hook_titulo": "La psicología del trader que pasa evaluaciones vs. el que las sigue comprando [hilo]",
        "copy": (
            "--- TWEET 1 (HOOK) ---\n"
            "He visto el patrón cientos de veces.\n\n"
            "Hay traders que pasan evaluaciones y hay traders que las siguen comprando.\n\n"
            "La diferencia no es la estrategia. Es la psicología.\n\n"
            "Hilo:\n\n"
            "--- TWEET 2 ---\n"
            "El trader que falla:\n\n"
            "→ Ajusta su estrategia después de cada pérdida\n"
            "→ Cree que el problema siempre es el setup\n"
            "→ No registra sus trades ni sus emociones\n"
            "→ Repite el mismo error sin saberlo\n\n"
            "--- TWEET 3 ---\n"
            "El trader que pasa:\n\n"
            "→ Tiene reglas claras. Las cumple aunque duela.\n"
            "→ Sabe que un día malo no define la evaluación\n"
            "→ Revisa su journal antes de operar, no después\n"
            "→ Reconoce cuándo está operando con ego, no con plan\n\n"
            "--- TWEET 4 ---\n"
            "El revenge trading es el enemigo silencioso.\n\n"
            "No lo sientes cuando sucede. Lo ves en los números después.\n\n"
            "Pierdes un trade → entras al siguiente con size más grande\n"
            "→ pierdes ese también → ya no es estrategia, es emociones.\n\n"
            "El journal existe para mostráartelo antes de que se convierta en un hábito.\n\n"
            "--- TWEET 5 ---\n"
            "La pregunta que deberías hacerte al final de cada sesión:\n\n"
            "'¿Cada trade de hoy tenía un plan antes de entrar?'\n\n"
            "Si la respuesta es no para más del 20% — tienes un problema de proceso, no de mercado.\n\n"
            "--- TWEET 6 ---\n"
            "Registrar tus trades con contexto emocional es la herramienta más subestimada en trading.\n\n"
            "No para 'sentirte mejor'. Para identificar cuándo tu PnL cae por mercado y cuándo cae por ti.\n\n"
            "Esa distinción vale más que cualquier indicador.\n\n"
            "¿Llevas journal emocional? ¿O solo registras entrada y salida?"
        ),
        "direccion_visual": (
            "Thread texto puro. Si se añade imagen al primer tweet: gráfico simple 'Trader que falla vs. trader que pasa' "
            "en dos columnas, fondo dark, tipografía blanca."
        ),
        "hashtags": "#propfirm #tradingpsychology #futurestrading #fundedtrader #FTMO #tradingjournal",
        "publicacion": "Lunes 08:30 AM ET",
        "hook_alternativo": "No es tu estrategia. Es tu psicología. Esto diferencia a los traders funded de los demás →",
        "notas_marca": "Thread de psicología — alto valor emocional para la audiencia. Diseñado para RT masivo.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 2 — Audiencia activa",
        "fecha": "2026-03-17",
        "plataforma": "X / Twitter",
        "pilar": "Producto en acción",
        "formato": "Tweet con imagen",
        "hook_titulo": "Tu equity curve dice cosas que tú no quieres ver",
        "copy": (
            "Tu equity curve dice cosas que tú no quieres ver.\n\n"
            "Esta es la diferencia entre un trader consistente y uno que vive de rachas:\n\n"
            "→ Curva consistente: sube con drawdowns pequeños y controlados\n"
            "→ Curva de racha: sube rápido, baja más rápido, nunca estable\n\n"
            "Las prop firms ven esto en segundos cuando evalúan tu cuenta.\n\n"
            "¿Cómo se ve la tuya?"
        ),
        "direccion_visual": (
            "Imagen: dos equity curves side by side en dark mode. "
            "Izquierda: curva suave ascendente (trader consistente), línea verde. "
            "Derecha: curva volátil con picos y valles (trader de rachas), línea amarilla/roja. "
            "Fondo #0D0D0D. Etiquetas simples. Opcional: marca de agua @zentrade_app."
        ),
        "hashtags": "#propfirm #futurestrading #tradingjournal #fundedtrader #FTMO",
        "publicacion": "Martes 09:00 AM ET",
        "hook_alternativo": "Las prop firms ven esto en tu cuenta en 10 segundos. ¿Sabes cómo se ve la tuya?",
        "notas_marca": "Visual educativo de alto impacto. La imagen hace el trabajo. El copy explica. Diseñar con cuidado.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 2 — Audiencia activa",
        "fecha": "2026-03-19",
        "plataforma": "X / Twitter",
        "pilar": "Educativo",
        "formato": "Tweet standalone",
        "hook_titulo": "El mejor setup del mundo no te salva si operas en el horario equivocado",
        "copy": (
            "El mejor setup del mundo no te salva si lo operas en el horario equivocado.\n\n"
            "La mayoría de traders de futuros tiene un 'horario dorado' donde su PnL es positivo — "
            "y un horario en el que sistemáticamente pierde.\n\n"
            "El problema: nadie lo mide.\n\n"
            "¿En qué horas del día tienes mejor win rate?"
        ),
        "direccion_visual": (
            "Tweet de texto puro. Sin imagen. Diseñado para engagement via replies."
        ),
        "hashtags": "#futurestrading #propfirm #tradingpsychology #fundedtrader",
        "publicacion": "Jueves 10:00 AM ET",
        "hook_alternativo": "Hay horas del día en las que eres un trader rentable. Otras en las que destruyes tu cuenta. ¿Sabes cuáles son?",
        "notas_marca": "Tweet de engagement. La pregunta final es el engagement hook. Responder con datos/análisis si hay replies.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 2 — Audiencia activa",
        "fecha": "2026-03-20",
        "plataforma": "X / Twitter",
        "pilar": "Comunidad",
        "formato": "Tweet pregunta",
        "hook_titulo": "¿Cuánto tiempo llevas sin pasar una evaluación?",
        "copy": (
            "Seamos honestos.\n\n"
            "¿Cuánto tiempo llevas intentando pasar tu primera (o siguiente) evaluación?\n\n"
            "A) Menos de 6 meses\n"
            "B) 6 a 12 meses\n"
            "C) Más de un año\n"
            "D) Ya tengo cuenta funded\n\n"
            "Sin juicio. Solo quiero entender dónde está la comunidad."
        ),
        "direccion_visual": "Tweet texto puro. Encuesta si la plataforma lo permite.",
        "hashtags": "#propfirm #fundedtrader #FTMO #futurestrading",
        "publicacion": "Viernes 01:00 PM ET",
        "hook_alternativo": "¿Cuánto llevas gastado en evaluaciones de prop firms? Comenta el número honesto.",
        "notas_marca": "Encuesta de comunidad. Responder a todos. Los datos de la encuesta pueden usarse en contenido futuro.",
        "estado": "Borrador",
    },

    # ---- Instagram ----
    {
        "semana": "Semana 2 — Audiencia activa",
        "fecha": "2026-03-17",
        "plataforma": "Instagram",
        "pilar": "Educativo",
        "formato": "Carrusel (6 slides)",
        "hook_titulo": "Qué revisan las prop firms en tu cuenta (que tú probablemente no estás midiendo)",
        "copy": (
            "--- SLIDE 1 ---\n"
            "Qué revisan las prop firms en tu cuenta\n"
            "(que tú probablemente no estás midiendo)\n\n"
            "--- SLIDE 2 ---\n"
            "01. Consistency score\n\n"
            "¿Ningún día representa más del 30-50% de tu ganancia total?\n"
            "Si lo hace, puedes fallar aunque hayas pasado el profit target.\n\n"
            "--- SLIDE 3 ---\n"
            "02. Equity curve\n\n"
            "No solo importa llegar al profit target.\n"
            "Importa CÓMO llegaste.\n"
            "Una curva suave y progresiva > un día de jackpot.\n\n"
            "--- SLIDE 4 ---\n"
            "03. Drawdown máximo usado\n\n"
            "¿Llegaste al 90% de tu límite de drawdown en algún momento?\n"
            "Para las prop firms eso es una bandera roja — aunque hayas recuperado.\n\n"
            "--- SLIDE 5 ---\n"
            "04. Días operados\n\n"
            "Muchas prop firms requieren un mínimo de días activos.\n"
            "No puedes hacer el profit target en 2 días y esperar que te aprueben.\n\n"
            "--- SLIDE 6 (CTA) ---\n"
            "¿Estás midiendo todo esto?\n\n"
            "Si tu respuesta es 'más o menos' — probablemente estás operando a ciegas.\n\n"
            "Próximamente: @zentrade_app calcula todo esto por ti, en tiempo real.\n\n"
            "Guarda este post para tu próxima evaluación."
        ),
        "direccion_visual": (
            "Fondo dark #0D0D0D. Numeración (01, 02...) en verde neón grande. "
            "Texto blanco, tipografía Inter o similar. "
            "Slide 6: fondo con degradado sutil oscuro y el logo de Zentrade centrado. "
            "Formato 1:1 cuadrado."
        ),
        "hashtags": "#propfirm #FTMO #ApexTrader #TopStep #futurestrading #tradingjournal #fundedtrader #tradingpsychology",
        "publicacion": "Martes 07:00 PM hora Colombia/México",
        "hook_alternativo": "Las prop firms evalúan esto en tu cuenta. La mayoría de traders no lo mide. ¿Y tú?",
        "notas_marca": "Carrusel educativo de alto valor. CTA suave — no vender, solo mencionar que viene la solución.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 2 — Audiencia activa",
        "fecha": "2026-03-19",
        "plataforma": "Instagram",
        "pilar": "Producto en acción",
        "formato": "Reel (45-60 seg)",
        "hook_titulo": "Así se ve el revenge trading en los datos (y por qué no lo ves mientras sucede)",
        "copy": (
            "--- GUIÓN REEL ---\n\n"
            "[00:00-00:03] HOOK VISUAL:\n"
            "Pantalla de dashboard con alerta naranja pulsante. "
            "Texto: 'Esto detectó tu journal después de tu peor sesión.'\n\n"
            "[00:03-00:20] NARRACIÓN:\n"
            "'El revenge trading no se siente como revenge trading cuando lo haces. "
            "Se siente como recuperar lo que perdiste. "
            "Pero en los datos se ve así: tres pérdidas consecutivas, "
            "siguiente entrada en menos de 30 minutos, tamaño de posición duplicado. "
            "Eso no es estrategia. Eso es emoción disfrazada de plan.'\n\n"
            "[00:20-00:40] SCREEN RECORDING:\n"
            "Mostrar la funcionalidad de detección de revenge trading del dashboard. "
            "Banner naranja con el patrón detectado. Animación suave.\n\n"
            "[00:40-00:55] CTA:\n"
            "'¿Cuántas veces te ha pasado esto sin saberlo? "
            "@zentrade_app lo detecta automáticamente. Lanzamos pronto.'\n\n"
            "--- CAPTION ---\n\n"
            "El revenge trading no se siente como revenge trading cuando lo haces.\n\n"
            "Se siente como recuperar lo que perdiste.\n\n"
            "Los datos cuentan una historia diferente.\n\n"
            "Síguenos @zentrade_app — la IA que detecta cuándo estás saboteando tu propia evaluación."
        ),
        "direccion_visual": (
            "Screen recording del dashboard con el banner de revenge trading visible. "
            "Overlay de texto en blanco/naranja sobre fondo oscuro. "
            "El color naranja del alerta debe ser protagonista — contrasta con el dark mode. "
            "Música: lo-fi instrumental con tensión leve. Formato 9:16 vertical."
        ),
        "hashtags": "#propfirm #tradingpsychology #futurestrading #fundedtrader #FTMO #tradingjournal #revengtrading",
        "publicacion": "Jueves 07:00 PM hora Colombia/México",
        "hook_alternativo": "¿Y si tu app te dijera cuándo estás haciendo revenge trading antes de que te vueles la cuenta?",
        "notas_marca": "Primera demostración del diferenciador #1 de Zentrade. Feature ZenMode. Mostrar sin explicar los precios.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 2 — Audiencia activa",
        "fecha": "2026-03-21",
        "plataforma": "Instagram",
        "pilar": "Behind the scenes",
        "formato": "Post imagen + caption",
        "hook_titulo": "Lo que nadie te dice sobre construir un SaaS mientras tradeas",
        "copy": (
            "Nadie te dice que construir un SaaS de trading mientras también tradeas "
            "te enseña más sobre psicología que cualquier curso.\n\n"
            "Cuando tienes una semana mala en el mercado y encima tienes que lanzar un feature nuevo... "
            "aprendes muy rápido la diferencia entre operar con datos y operar con emociones.\n\n"
            "Eso es exactamente lo que quisimos resolver con Zentrade.\n\n"
            "No solo un journal. Una herramienta que te muestra cuándo eres tú el problema "
            "— y cuándo es el mercado.\n\n"
            "Lanzamos pronto. Síguenos @zentrade_app."
        ),
        "direccion_visual": (
            "Imagen: pantalla dividida — lado izquierdo un gráfico de trading, lado derecho código o dashboard de Zentrade. "
            "Estética dark mode en ambos lados. "
            "Alternativa: foto del fundador trabajando en laptop con setup de trading de fondo. "
            "Formato 1:1."
        ),
        "hashtags": "#buildinpublic #propfirm #saas #futurestrading #tradingjournal #indiedev",
        "publicacion": "Sábado 11:00 AM hora Colombia/México",
        "hook_alternativo": "Ser developer y trader al mismo tiempo me enseñó algo que ningún curso te dice.",
        "notas_marca": "Post de conexión. Humaniza la marca. No hay CTA de ventas — solo seguir la cuenta.",
        "estado": "Borrador",
    },

    # ---- TikTok ----
    {
        "semana": "Semana 2 — Audiencia activa",
        "fecha": "2026-03-18",
        "plataforma": "TikTok",
        "pilar": "Educativo",
        "formato": "Video (60 seg)",
        "hook_titulo": "Por qué tu win rate no significa nada sin este otro número",
        "copy": (
            "--- GUIÓN TIKTOK ---\n\n"
            "[00:00-00:03] HOOK:\n"
            "'Un win rate del 70% puede hacerte perder dinero. Te explico cómo.'\n\n"
            "[00:03-00:25] DESARROLLO:\n"
            "'Imagina que ganas 7 de cada 10 trades. Suena increíble, ¿verdad? "
            "Pero si cada trade ganador te da 0.5 RR y cada perdedor te cuesta 1 RR... "
            "7 × 0.5 = 3.5 puntos ganados. 3 × 1 = 3 puntos perdidos. "
            "Eso es 0.5 puntos de ventaja. Breakeven en el mejor caso cuando pones comisiones.'\n\n"
            "[00:25-00:45] SOLUCIÓN:\n"
            "'Lo que necesitas es profit factor. "
            "Ganancia bruta dividido pérdida bruta. "
            "Si es mayor a 1.5, tienes una estrategia real. "
            "Si es menor a 1.0, estás perdiendo dinero aunque tengas win rate alto.'\n\n"
            "[00:45-00:60] CTA:\n"
            "'¿Sabes cuál es tu profit factor ahora mismo? "
            "Comenta abajo — y sígueme para más métricas que los traders funded sí miden.'\n\n"
            "--- TEXTO EN PANTALLA ---\n"
            "0:00 → 'Win rate 70% = pérdida de dinero?'\n"
            "0:15 → '7 × 0.5 RR - 3 × 1 RR = 0.5 (sin comisiones)'\n"
            "0:30 → 'Profit Factor = Ganancia bruta ÷ Pérdida bruta'\n"
            "0:45 → '>1.5 = estrategia real | <1.0 = pierdes dinero'"
        ),
        "direccion_visual": (
            "Talking head a cámara O screen recording con voz en off. "
            "Texto superpuesto con los números/fórmulas — esencial para que funcione sin audio. "
            "Fondo dark si es talking head. "
            "Captions automáticos activados."
        ),
        "hashtags": "#propfirm #futurestrading #tradingpsychology #fundedtrader #FTMO #tradertips #profitfactor",
        "publicacion": "Miércoles 06:00 PM hora Colombia/México",
        "hook_alternativo": "El número que los traders funded miran más que el win rate (y que nadie te enseña)",
        "notas_marca": "Video educativo con math visible. El contenido con números/fórmulas genera guardados y compartidos.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 2 — Audiencia activa",
        "fecha": "2026-03-21",
        "plataforma": "TikTok",
        "pilar": "Producto en acción",
        "formato": "Video screen recording (30-45 seg)",
        "hook_titulo": "El dashboard que quisiera haber tenido en mi primera evaluación",
        "copy": (
            "--- GUIÓN TIKTOK ---\n\n"
            "[00:00-00:03] HOOK (voz en off):\n"
            "'Este dashboard me habría salvado al menos 3 evaluaciones fallidas.'\n\n"
            "[00:03-00:30] SCREEN RECORDING CON VOZ:\n"
            "'Miren: profit factor, consistency score, equity curve, todos en tiempo real. "
            "¿Ven este número acá? Es el porcentaje de consistencia. "
            "Si llega al 45%, la evaluación está en riesgo — sin importar cuánto profit tenga. "
            "Exactamente lo que ningún otro journal mide de esta forma para prop firms.'\n\n"
            "[00:30-00:43] CTA:\n"
            "'@zentrade_app — lanzamos pronto. Link en bio para ser de los primeros.'\n\n"
            "--- TEXTO SUPERPUESTO ---\n"
            "0:00 → '3 evaluaciones fallidas. Este dashboard las habría evitado.'\n"
            "0:10 → 'Consistency score en tiempo real'\n"
            "0:35 → 'Link en bio → acceso anticipado'"
        ),
        "direccion_visual": (
            "Screen recording directo del dashboard de Zentrade con overlay de voz. "
            "Zoom suave sobre los KPIs clave. "
            "Cursor que guía la atención. Datos demo reales y convincentes. "
            "Formato vertical 9:16."
        ),
        "hashtags": "#propfirm #tradingjournal #futurestrading #fundedtrader #FTMO #TopStep #tradingapp",
        "publicacion": "Sábado 12:00 PM hora Colombia/México",
        "hook_alternativo": "Así se ve el journal que los traders que pasan evaluaciones necesitan (y no existía en español)",
        "notas_marca": "Demo de producto sin vender. Dejar que el producto hable. Dirección visual > copy en este post.",
        "estado": "Borrador",
    },

    # =========================================================
    # SEMANA 3 — ACTIVACIÓN DE INFLUENCERS
    # =========================================================

    # ---- X / Twitter ----
    {
        "semana": "Semana 3 — Influencers",
        "fecha": "2026-03-23",
        "plataforma": "X / Twitter",
        "pilar": "Educativo",
        "formato": "Thread (5 tweets)",
        "hook_titulo": "Cómo analizo mi semana de trading en 10 minutos (y por qué la mayoría no lo hace) [hilo]",
        "copy": (
            "--- TWEET 1 ---\n"
            "Cómo analizo mi semana de trading en 10 minutos.\n\n"
            "Y por qué la mayoría de traders no lo hace.\n\n"
            "Hilo:\n\n"
            "--- TWEET 2 ---\n"
            "Pregunta 1: ¿Cuántos trades tomé que NO tenían un plan antes de entrar?\n\n"
            "Si es más del 20%, esa semana no fue una semana de trading.\n"
            "Fue una semana de gambling con buena presentación.\n\n"
            "--- TWEET 3 ---\n"
            "Pregunta 2: ¿En qué horarios concentré mis pérdidas?\n\n"
            "No para evitar esas horas — sino para entender si el mercado cambia "
            "o si YO cambio en esas horas.\n\n"
            "Son cosas muy diferentes.\n\n"
            "--- TWEET 4 ---\n"
            "Pregunta 3: ¿Hubo algún momento en que entré a un trade para 'recuperar'?\n\n"
            "Si la respuesta es sí — ese trade no cuenta como trade. "
            "Cuenta como reacción emocional disfrazada de setup.\n\n"
            "--- TWEET 5 ---\n"
            "Pregunta 4: ¿Mi consistency score está dentro del rango que requiere mi prop firm?\n\n"
            "Esta pregunta sola puede salvarte de una evaluación fallida.\n\n"
            "--- TWEET 6 ---\n"
            "Estas preguntas toman 10 minutos.\n\n"
            "Pero solo funcionan si tienes los datos.\n\n"
            "Un journal no es un diario. Es tu sistema de retroalimentación.\n\n"
            "¿Qué pregunta te parece más difícil de responder honestamente?"
        ),
        "direccion_visual": "Thread texto puro. Sin imagen.",
        "hashtags": "#tradingjournal #propfirm #futurestrading #tradingpsychology #fundedtrader",
        "publicacion": "Lunes 09:00 AM ET",
        "hook_alternativo": "4 preguntas que los traders funded se hacen cada domingo. ¿Te las estás haciendo tú?",
        "notas_marca": "Thread de valor alto orientado a hábitos. Genera bookmarks. No mencionar Zentrade hasta el final y de forma sutil.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 3 — Influencers",
        "fecha": "2026-03-25",
        "plataforma": "X / Twitter",
        "pilar": "Producto en acción",
        "formato": "Tweet con imagen/video",
        "hook_titulo": "El reporte semanal que Zentrade genera automáticamente",
        "copy": (
            "Todos los lunes, Zentrade te manda esto:\n\n"
            "→ Resumen de la semana con tus KPIs reales\n"
            "→ Análisis de tus patrones emocionales\n"
            "→ Detección de momentos de revenge trading\n"
            "→ Recomendaciones específicas para la semana siguiente\n\n"
            "Generado por IA. Enviado a tu email. Listo antes de que abras el mercado.\n\n"
            "Lanzamos pronto. Sigue @zentrade_app."
        ),
        "direccion_visual": (
            "Screenshot del email de reporte semanal (dark mode, bien diseñado). "
            "Si no hay screenshot disponible: mockup de email en phone/desktop en dark mode. "
            "Texto del tweet como preview del valor."
        ),
        "hashtags": "#propfirm #tradingjournal #futurestrading #fundedtrader #FTMO #AI",
        "publicacion": "Miércoles 09:00 AM ET",
        "hook_alternativo": "Imagina recibir este análisis cada lunes antes de operar →",
        "notas_marca": "Feature showcase del reporte semanal (Pro/ZenMode). Mostrar el email real si está disponible.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 3 — Influencers",
        "fecha": "2026-03-27",
        "plataforma": "X / Twitter",
        "pilar": "Educativo",
        "formato": "Tweet standalone",
        "hook_titulo": "El error más caro en Apex Trader: confundir trailing drawdown con drawdown máximo",
        "copy": (
            "El error más caro en Apex Trader:\n\n"
            "Confundir trailing drawdown con drawdown máximo.\n\n"
            "Trailing drawdown = baja cuando pierdes, pero NUNCA SUBE cuando ganas.\n"
            "Drawdown máximo = un techo fijo desde el balance inicial.\n\n"
            "En Apex, tu trailing drawdown sigue tu equity máxima.\n"
            "Si llegas a $53,000 y bajas a $51,500 → tu floor ahora es $51,500, no $48,000.\n\n"
            "No confundirlo puede salvarte una cuenta.\n\n"
            "¿Estabas al tanto de esta diferencia?"
        ),
        "direccion_visual": (
            "Imagen explicativa: diagrama simple de trailing drawdown vs. max drawdown. "
            "Línea de equity subiendo, floor del trailing drawdown subiendo con ella. "
            "Fondo dark. Colores: equity en verde, floor en rojo, zona de riesgo sombreada."
        ),
        "hashtags": "#ApexTrader #propfirm #futurestrading #fundedtrader #tradingjournal",
        "publicacion": "Viernes 09:00 AM ET",
        "hook_alternativo": "Apex Trader tiene una regla que destruye cuentas porque nadie la entiende bien. Esta es.",
        "notas_marca": "Post técnico específico de Apex. Alta relevancia para la audiencia objetivo. Posiciona autoridad.",
        "estado": "Borrador",
    },

    # ---- Instagram ----
    {
        "semana": "Semana 3 — Influencers",
        "fecha": "2026-03-24",
        "plataforma": "Instagram",
        "pilar": "Educativo",
        "formato": "Carrusel (5 slides)",
        "hook_titulo": "La rutina pre-trading de los traders que pasan evaluaciones",
        "copy": (
            "--- SLIDE 1 ---\n"
            "La rutina pre-trading de los traders que pasan evaluaciones\n\n"
            "--- SLIDE 2 ---\n"
            "30 min antes de abrir el mercado:\n\n"
            "Revisar journal del día anterior.\n"
            "¿Hubo revenge trading? ¿Saliste del plan? ¿Por qué?\n\n"
            "--- SLIDE 3 ---\n"
            "15 min antes:\n\n"
            "Revisar KPIs de la evaluación.\n"
            "¿Cuánto drawdown llevo? ¿Cuál es mi consistency score actual?\n"
            "¿Puedo operar hoy con tamaño normal o debo reducir?\n\n"
            "--- SLIDE 4 ---\n"
            "Justo antes de abrir:\n\n"
            "Definir máximo de pérdida del día.\n"
            "No el de la prop firm — el tuyo propio. Más conservador.\n"
            "Si lo tocas: cierra. Sin negociar.\n\n"
            "--- SLIDE 5 ---\n"
            "Durante la sesión:\n\n"
            "Registrar cada trade con el setup y el estado emocional.\n"
            "No para el historial. Para leerlo 10 minutos después de cerrar.\n\n"
            "--- SLIDE 6 (CTA) ---\n"
            "La diferencia no está en el setup.\n"
            "Está en el proceso antes y después de operar.\n\n"
            "Guarda este post para tu próxima sesión.\n"
            "@zentrade_app — lanzamos pronto."
        ),
        "direccion_visual": (
            "Fondo dark. Icono de reloj o línea de tiempo vertical en verde neón. "
            "Tipografía limpia, texto centrado. "
            "Cada slide con un horario visible prominente (30 min, 15 min, etc). "
            "Slide final con logo Zentrade y degradado oscuro de fondo."
        ),
        "hashtags": "#propfirm #tradingpsychology #futurestrading #fundedtrader #FTMO #tradingjournal #tradinghabits",
        "publicacion": "Martes 07:00 PM hora Colombia/México",
        "hook_alternativo": "Lo que los traders funded hacen ANTES de operar (que nadie más hace)",
        "notas_marca": "Carrusel de rutinas — alto valor práctico, alta tasa de guardado. Terminar con CTA suave.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 3 — Influencers",
        "fecha": "2026-03-26",
        "plataforma": "Instagram",
        "pilar": "Comunidad",
        "formato": "Post imagen + caption",
        "hook_titulo": "¿Cuál es tu prop firm y por qué la elegiste?",
        "copy": (
            "Si tuvieras que recomendar UNA prop firm para un trader de futuros que está empezando...\n\n"
            "¿Cuál sería y por qué?\n\n"
            "Nosotros tenemos una opinión pero primero queremos escuchar la tuya.\n\n"
            "Comenta abajo. Leemos todo.\n\n"
            "@zentrade_app"
        ),
        "direccion_visual": (
            "Imagen con logos de las principales prop firms (FTMO, Apex, TopStep, Tradoverse, Uprofit) "
            "en estilo dark mode, ordenados visualmente. "
            "Fondo #0D0D0D. Logos en blanco/gris. "
            "Texto superpuesto: '¿Cuál es la tuya?'"
        ),
        "hashtags": "#propfirm #FTMO #ApexTrader #TopStep #Tradovate #Uprofit #futurestrading #fundedtrader",
        "publicacion": "Jueves 06:00 PM hora Colombia/México",
        "hook_alternativo": "FTMO, Apex, TopStep, Tradoverse... ¿cuál recomiendas y por qué?",
        "notas_marca": "Post de engagement de comunidad. Responder TODOS los comentarios. Los datos pueden informar contenido futuro.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 3 — Influencers",
        "fecha": "2026-03-28",
        "plataforma": "Instagram",
        "pilar": "Producto en acción",
        "formato": "Reel (45-60 seg)",
        "hook_titulo": "Así se ve un reporte semanal generado por IA para traders",
        "copy": (
            "--- GUIÓN REEL ---\n\n"
            "[00:00-00:03] HOOK:\n"
            "Pantalla de email abriéndose. Texto: 'Este email llega todos los lunes.'\n\n"
            "[00:03-00:30] DEMO:\n"
            "Screen recording del email de reporte semanal. Narración:\n"
            "'Imagina recibir esto antes de abrir el mercado cada semana: "
            "tu profit factor de los últimos 5 días, tus horas más rentables, "
            "y un análisis de cuándo operaste con emoción y cuándo con plan. "
            "Generado por IA con tus datos reales.'\n\n"
            "[00:30-00:50] IMPACTO:\n"
            "'No para que te sientas bien. Para que el lunes tomes mejores decisiones que el viernes pasado.'\n\n"
            "[00:50-00:60] CTA:\n"
            "'@zentrade_app — acceso anticipado en el link de bio.'\n\n"
            "--- CAPTION ---\n\n"
            "Todos los lunes. Este reporte. Tu semana de trading, analizada por IA.\n\n"
            "Patrones emocionales. Revenge trading. Horas de mejor rendimiento. Todo en un email.\n\n"
            "Lanzamos pronto. Acceso anticipado en el link de bio.\n"
            "@zentrade_app"
        ),
        "direccion_visual": (
            "Screen recording del email en formato mobile (inbox → abrir email → scroll). "
            "Dark mode del email visible. "
            "Música: ambient/lo-fi profesional. "
            "Texto overlay en blanco sobre partes clave del email."
        ),
        "hashtags": "#propfirm #tradingjournal #futurestrading #fundedtrader #AI #tradingpsychology #FTMO",
        "publicacion": "Sábado 11:00 AM hora Colombia/México",
        "hook_alternativo": "¿Y si recibieras este análisis de tu semana de trading cada lunes antes de abrir el mercado?",
        "notas_marca": "Demo del reporte semanal (feature Pro/ZenMode). Mostrar el email real si está disponible.",
        "estado": "Borrador",
    },

    # ---- TikTok ----
    {
        "semana": "Semana 3 — Influencers",
        "fecha": "2026-03-24",
        "plataforma": "TikTok",
        "pilar": "Educativo",
        "formato": "Video (60 seg)",
        "hook_titulo": "Por qué el 90% de traders que pasa la fase 1 falla la fase 2",
        "copy": (
            "--- GUIÓN TIKTOK ---\n\n"
            "[00:00-00:03] HOOK:\n"
            "'Pasaste la fase 1 de tu evaluación. Felicidades. Ahora viene la parte donde el 90% falla.'\n\n"
            "[00:03-00:25] DESARROLLO:\n"
            "'La fase 2 no es más difícil porque el mercado cambia. "
            "Es más difícil porque TÚ cambias. "
            "Ahora operas con presión. Con miedo a fallar lo que ya conseguiste. "
            "Y esa presión hace que sobreoperes, que reduzcas size cuando no deberías, "
            "o que abandones tu estrategia justo cuando más la necesitas.'\n\n"
            "[00:25-00:50] SOLUCIÓN:\n"
            "'La respuesta no es 'ser más disciplinado'. "
            "La respuesta es tener datos. "
            "Si sabes que tus mejores semanas son cuando operas 2-3 trades al día, "
            "en las primeras 2 horas del mercado, con stop de 1.5 puntos — "
            "no hay nada que la presión te pueda quitar. Los números lo saben por ti.'\n\n"
            "[00:50-00:60] CTA:\n"
            "'Sígueme. Más sobre psicología y métricas de evaluación cada semana.'\n\n"
            "--- TEXTO PANTALLA ---\n"
            "0:00 → 'Por qué fallas la fase 2 (aunque pasaste la 1)'\n"
            "0:15 → 'No es el mercado. Eres tú bajo presión.'\n"
            "0:40 → 'Los datos te quitan la presión de encima'"
        ),
        "direccion_visual": (
            "Talking head a cámara con fondo dark (setup de trading). "
            "Texto superpuesto con puntos clave. "
            "Energía: directa, sin rodeos, empática."
        ),
        "hashtags": "#propfirm #FTMO #futurestrading #fundedtrader #tradingpsychology #evaluacion",
        "publicacion": "Martes 07:00 PM hora Colombia/México",
        "hook_alternativo": "Pasaste la fase 1. Aquí está por qué la mayoría no pasa la fase 2.",
        "notas_marca": "Contenido de psicología que resuena profundo en la audiencia. Alta probabilidad de comentarios y shares.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 3 — Influencers",
        "fecha": "2026-03-27",
        "plataforma": "TikTok",
        "pilar": "Behind the scenes",
        "formato": "Video (30-45 seg)",
        "hook_titulo": "Semana 3 construyendo en público: lo que nadie te muestra del pre-lanzamiento",
        "copy": (
            "--- GUIÓN TIKTOK ---\n\n"
            "[00:00-00:03] HOOK:\n"
            "'Semana 3 construyendo @zentrade_app en público. Aquí va la realidad.'\n\n"
            "[00:03-00:30] BUILD IN PUBLIC:\n"
            "'Esta semana terminamos el sistema de reporte semanal con IA. "
            "También activamos el detector de revenge trading — la primera vez que lo probé con datos reales "
            "me mostró exactamente lo que no quería ver sobre mi propia forma de operar. "
            "Eso es lo que queremos que haga con ustedes.'\n\n"
            "[00:30-00:43] CTA:\n"
            "'Lanzamos en semanas. Si quieres ser de los primeros — link en bio. "
            "Comenta qué feature quieren ver primero.'\n\n"
            "--- TEXTO PANTALLA ---\n"
            "0:00 → 'Semana 3 — build in public'\n"
            "0:15 → 'Reporte semanal IA + detector de revenge trading'\n"
            "0:35 → 'Lanzamos pronto — link en bio'"
        ),
        "direccion_visual": (
            "Selfie a cámara con pantalla de código/dashboard de fondo. "
            "Muy auténtico — sin producción alta. "
            "Alternativa: split screen (código + dashboard resultado). "
            "Energía: entusiasta pero sin exagerar."
        ),
        "hashtags": "#buildinpublic #saas #propfirm #futurestrading #tradingjournal #indiedev #fundedtrader",
        "publicacion": "Viernes 07:00 PM hora Colombia/México",
        "hook_alternativo": "Lo que realmente pasa 2 semanas antes del lanzamiento de un SaaS de trading",
        "notas_marca": "Build in public Semana 3. Consistencia en esta narrativa semanal crea audiencia fiel.",
        "estado": "Borrador",
    },

    # =========================================================
    # SEMANA 4 — PRE-LAUNCH MOMENTUM
    # =========================================================

    # ---- X / Twitter ----
    {
        "semana": "Semana 4 — Pre-launch",
        "fecha": "2026-03-30",
        "plataforma": "X / Twitter",
        "pilar": "Producto en acción",
        "formato": "Tweet standalone (anuncio)",
        "hook_titulo": "Lanzamos en días — 50 primeros usuarios reciben 30 días Pro gratis",
        "copy": (
            "Lanzamos @zentrade_app en días.\n\n"
            "Los 50 primeros usuarios registrados reciben 30 días de Plan Professional gratis. Sin tarjeta. Sin trampa.\n\n"
            "Lo que incluye:\n"
            "→ Dashboard analítico completo\n"
            "→ Reporte semanal con IA\n"
            "→ Equity curve + profit factor en tiempo real\n"
            "→ Benchmarking vs. reglas de tu prop firm\n\n"
            "El link estará en bio cuando abramos.\n\n"
            "RT si conoces a alguien en evaluación que lo necesita."
        ),
        "direccion_visual": (
            "Tweet texto puro con RT como CTA principal. "
            "Opcional: imagen con '50 accesos anticipados — 30 días Pro gratis' "
            "en tipografía grande sobre fondo dark con acento verde."
        ),
        "hashtags": "#propfirm #futurestrading #fundedtrader #FTMO #TopStep #ApexTrader #tradingjournal",
        "publicacion": "Lunes 09:00 AM ET",
        "hook_alternativo": "50 plazas. 30 días Pro gratis. Lanzamos esta semana. RT para que lo vea quien lo necesita.",
        "notas_marca": "SCARCITY + RECIPROCITY. El primer post que menciona la oferta de lanzamiento. Pinear en el perfil.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 4 — Pre-launch",
        "fecha": "2026-04-01",
        "plataforma": "X / Twitter",
        "pilar": "Behind the scenes",
        "formato": "Thread — build story (7 tweets)",
        "hook_titulo": "Cómo construimos Zentrade en X meses — la historia completa [hilo]",
        "copy": (
            "--- TWEET 1 ---\n"
            "Hoy cuento cómo construimos @zentrade_app desde cero.\n\n"
            "El producto, los errores, las decisiones. Sin filtro.\n\n"
            "Hilo:\n\n"
            "--- TWEET 2 ---\n"
            "Todo empezó con una pregunta simple:\n\n"
            "'¿Por qué no existe un journal de trading diseñado específicamente para traders en evaluación de prop firms?'\n\n"
            "Probé TradeZella, Edgewonk, Tradervue.\n"
            "Todos buenos. Ninguno pensado para lo que las prop firms realmente miden.\n"
            "Ninguno en español. Ninguno con IA orientada a psicología de evaluación.\n\n"
            "--- TWEET 3 ---\n"
            "El primer prototipo era feo y funcional.\n\n"
            "Un dashboard en Next.js con las 5 métricas que más importan en evaluación:\n"
            "consistency %, profit factor, equity curve, drawdown relativo, y días operados.\n\n"
            "Sin diseño. Sin pulir. Solo los números que necesitaba.\n\n"
            "--- TWEET 4 ---\n"
            "El momento en que supe que valía la pena:\n\n"
            "Conecté mis trades reales. El dashboard me mostró que el 60% de mis pérdidas "
            "ocurrían en los primeros 30 minutos después de perder un trade.\n\n"
            "Revenge trading. Lo sabía. Pero nunca lo había visto en datos.\n\n"
            "--- TWEET 5 ---\n"
            "Entonces construimos el detector de revenge trading.\n\n"
            "Un algoritmo que identifica cuándo entraste a un trade en menos de 30 minutos "
            "después de una pérdida, o cuando llevabas 3+ pérdidas consecutivas.\n\n"
            "No para juzgarte. Para mostrártelo antes de que se convierta en un hábito.\n\n"
            "--- TWEET 6 ---\n"
            "También añadimos el reporte semanal con IA.\n\n"
            "Todos los lunes, un análisis completo de tu semana: KPIs, patrones emocionales, "
            "recomendaciones específicas para la siguiente semana.\n\n"
            "Generado por Gemini. Enviado a tu email antes de que abras el mercado.\n\n"
            "--- TWEET 7 ---\n"
            "Hoy lanzamos.\n\n"
            "Los 50 primeros usuarios reciben 30 días de Plan Professional gratis.\n\n"
            "Si estás en evaluación de prop firm, si llevas tiempo buscando un journal que te dé "
            "datos reales sobre tus patrones — esto es para ti.\n\n"
            "Link en bio → @zentrade_app\n\n"
            "Y si llegaste hasta acá: gracias. Esto lo construimos para traders como tú."
        ),
        "direccion_visual": (
            "Thread texto puro. Altamente conversacional y personal. "
            "El último tweet puede tener screenshot del dashboard como imagen de cierre."
        ),
        "hashtags": "#buildinpublic #propfirm #futurestrading #saas #fundedtrader #tradingjournal #FTMO",
        "publicacion": "Miércoles 09:00 AM ET",
        "hook_alternativo": "El hilo de cómo nació Zentrade — desde la primera idea hasta el lanzamiento de hoy →",
        "notas_marca": "El thread más importante de toda la campaña. La narrativa de origen. Pinear. Compartir en todas las plataformas.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 4 — Pre-launch",
        "fecha": "2026-04-03",
        "plataforma": "X / Twitter",
        "pilar": "Comunidad",
        "formato": "Tweet standalone",
        "hook_titulo": "Lanzamos hoy. Gracias a los que estuvieron desde el principio.",
        "copy": (
            "Hoy lanzamos @zentrade_app.\n\n"
            "Gracias a los que siguieron, los que comentaron, los que hicieron RT desde la semana 1.\n\n"
            "Construimos esto para traders de futuros que van en serio con las prop firms.\n"
            "Para traders que quieren datos, no corazonadas.\n"
            "Para traders cansados de repetir los mismos errores sin entender por qué.\n\n"
            "Link en bio → 50 plazas con 30 días Pro gratis.\n\n"
            "Hoy empieza."
        ),
        "direccion_visual": "Tweet texto puro. Máximo de autenticidad.",
        "hashtags": "#propfirm #futurestrading #fundedtrader #tradingjournal #FTMO #lanzamiento",
        "publicacion": "Viernes 09:00 AM ET",
        "hook_alternativo": "Lanzamos. 50 plazas. 30 días Pro gratis. Esto es para los que van en serio.",
        "notas_marca": "Tweet de lanzamiento final. Emocional y directo. Máxima autenticidad.",
        "estado": "Borrador",
    },

    # ---- Instagram ----
    {
        "semana": "Semana 4 — Pre-launch",
        "fecha": "2026-03-31",
        "plataforma": "Instagram",
        "pilar": "Producto en acción",
        "formato": "Carrusel (6 slides)",
        "hook_titulo": "Todo lo que incluye Zentrade (y por qué lo construimos así)",
        "copy": (
            "--- SLIDE 1 ---\n"
            "Zentrade — el journal para traders que van en serio con las prop firms\n\n"
            "--- SLIDE 2 ---\n"
            "Dashboard KPI en tiempo real\n\n"
            "Profit factor, consistency score, equity curve.\n"
            "Todos los números que las prop firms miden — visibles antes de operar.\n\n"
            "--- SLIDE 3 ---\n"
            "Detección de revenge trading con IA\n\n"
            "El algoritmo identifica cuándo estás operando con emoción, no con plan.\n"
            "Un banner de alerta antes de que lo repitas.\n\n"
            "--- SLIDE 4 ---\n"
            "Reporte semanal inteligente\n\n"
            "Todos los lunes. Tu semana analizada: KPIs, patrones emocionales, recomendaciones.\n"
            "Generado por IA. Enviado a tu email.\n\n"
            "--- SLIDE 5 ---\n"
            "Calendario emocional + journal diario\n\n"
            "Registra cómo te sentiste cada sesión. Correlaciona emociones con PnL.\n"
            "Ve exactamente cuándo el mercado te gana y cuándo te ganas tú solo.\n\n"
            "--- SLIDE 6 (CTA — SCARCITY) ---\n"
            "Lanzamos esta semana.\n\n"
            "Los 50 primeros usuarios registrados reciben 30 días de Plan Professional gratis.\n\n"
            "Link en bio → @zentrade_app\n\n"
            "Comparte esto con el trader que lo necesita."
        ),
        "direccion_visual": (
            "Carrusel de producto. Cada slide muestra un screenshot real de la feature + descripción. "
            "Fondo dark #0D0D0D. Screenshots del dashboard con bordes redondeados y sombra sutil. "
            "Slide 1: solo tipografía y logo, impactante. "
            "Slide 6: fondo degradado oscuro → verde con el número '50' prominente."
        ),
        "hashtags": "#propfirm #FTMO #TopStep #ApexTrader #futurestrading #tradingjournal #fundedtrader #tradingpsychology #NinjaTrader #Tradovate",
        "publicacion": "Martes 07:00 PM hora Colombia/México",
        "hook_alternativo": "Construimos la herramienta que los traders de prop firms no tenían. Esto es Zentrade.",
        "notas_marca": "Carrusel de lanzamiento completo. El más importante de Instagram. Máxima calidad de diseño.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 4 — Pre-launch",
        "fecha": "2026-04-02",
        "plataforma": "Instagram",
        "pilar": "Producto en acción",
        "formato": "Reel (60 seg) — lanzamiento",
        "hook_titulo": "Zentrade ya está disponible — tour completo del dashboard",
        "copy": (
            "--- GUIÓN REEL LANZAMIENTO ---\n\n"
            "[00:00-00:03] HOOK:\n"
            "Dashboard aparece en pantalla. Música con energía. "
            "Texto: 'Ya está disponible. Así se ve por dentro.'\n\n"
            "[00:03-00:45] TOUR DEL PRODUCTO:\n"
            "Screen recording fluido a través del dashboard:\n"
            "1. Dashboard principal — KPIs y equity curve\n"
            "2. Calendario con días verdes/rojos\n"
            "3. Detección de revenge trading (banner naranja)\n"
            "4. Preview del reporte semanal\n\n"
            "Narración: 'Esto es Zentrade. Un journal de trading construido para traders que van en serio "
            "con las prop firms. Con IA que detecta cuándo te estás saboteando. "
            "Con métricas que las prop firms realmente miden. En español, para LATAM.'\n\n"
            "[00:45-00:60] CTA:\n"
            "'Los primeros 50 usuarios reciben 30 días Pro gratis. Sin tarjeta. Link en bio.'\n\n"
            "--- CAPTION ---\n\n"
            "Ya lanzamos.\n\n"
            "Zentrade: el journal de trading con IA para traders en evaluación de prop firms.\n\n"
            "Los primeros 50 usuarios registrados reciben 30 dias de Plan Professional gratis. Sin tarjeta. Sin trampa.\n\n"
            "Link en bio. Comparte con quien lo necesite.\n\n"
            "@zentrade_app"
        ),
        "direccion_visual": (
            "Screen recording del dashboard a 60fps si es posible. "
            "Música energizante pero profesional — sin trap. "
            "Transitions suaves entre secciones. "
            "Texto overlay minimalista. "
            "Último frame: logo Zentrade + URL + '50 plazas disponibles'."
        ),
        "hashtags": "#propfirm #FTMO #TopStep #ApexTrader #futurestrading #tradingjournal #fundedtrader #tradingapp #lanzamiento",
        "publicacion": "Jueves 07:00 PM hora Colombia/México",
        "hook_alternativo": "Esto tardó meses en construirse. Ya está disponible. Aquí el tour completo →",
        "notas_marca": "Reel de lanzamiento. El más importante del canal. Máxima producción. Impulsar con pauta si hay presupuesto.",
        "estado": "Borrador",
    },

    # ---- TikTok ----
    {
        "semana": "Semana 4 — Pre-launch",
        "fecha": "2026-03-30",
        "plataforma": "TikTok",
        "pilar": "Producto en acción",
        "formato": "Video (45-60 seg)",
        "hook_titulo": "50 traders van a recibir esto gratis esta semana",
        "copy": (
            "--- GUIÓN TIKTOK ---\n\n"
            "[00:00-00:03] HOOK:\n"
            "'50 traders van a recibir acceso Pro a Zentrade completamente gratis esta semana.'\n\n"
            "[00:03-00:30] PROPUESTA DE VALOR:\n"
            "'Zentrade es el primer journal de trading en español diseñado para prop firms. "
            "Dashboard con profit factor, consistency score, equity curve en tiempo real. "
            "Detección de revenge trading con IA. "
            "Reporte semanal generado por inteligencia artificial con tus propios datos.'\n\n"
            "[00:30-00:50] OFERTA + URGENCIA:\n"
            "'Los primeros 50 que se registren reciben 30 días del plan Professional. "
            "Gratis. Sin tarjeta. Porque queremos que lo prueben traders reales en evaluaciones reales.'\n\n"
            "[00:50-00:60] CTA:\n"
            "'Link en bio. Comenta si tienes alguna pregunta — respondo todo.'\n\n"
            "--- TEXTO PANTALLA ---\n"
            "0:00 → '50 accesos Pro GRATIS — esta semana'\n"
            "0:15 → 'Journal para prop firms con IA'\n"
            "0:35 → '30 días Pro | sin tarjeta | en español'\n"
            "0:50 → 'Link en bio → @zentrade_app'"
        ),
        "direccion_visual": (
            "Screen recording del dashboard con overlay de voz. "
            "Texto superpuesto destacando los números (50, 30 días, gratis). "
            "Energía: directa y con urgencia, no exagerada. "
            "Formato 9:16."
        ),
        "hashtags": "#propfirm #FTMO #ApexTrader #TopStep #futurestrading #fundedtrader #tradingjournal #gratis",
        "publicacion": "Lunes 07:00 PM hora Colombia/México",
        "hook_alternativo": "Esta semana, 50 traders van a probar gratis lo que los traders funded usan para no sabotearse.",
        "notas_marca": "Video de scarcity + lanzamiento. Responder comentarios con urgencia. Actualizar si se agotan las plazas.",
        "estado": "Borrador",
    },
    {
        "semana": "Semana 4 — Pre-launch",
        "fecha": "2026-04-02",
        "plataforma": "TikTok",
        "pilar": "Behind the scenes",
        "formato": "Video (30-45 seg)",
        "hook_titulo": "El día del lanzamiento — lo que nadie muestra",
        "copy": (
            "--- GUIÓN TIKTOK ---\n\n"
            "[00:00-00:03] HOOK:\n"
            "'Hoy lanzamos @zentrade_app. Esto es lo que nadie te muestra del día del lanzamiento.'\n\n"
            "[00:03-00:30] AUTENTICIDAD:\n"
            "'Son las 8 de la mañana. El servidor está corriendo. "
            "El link ya está activo. "
            "Llevamos meses construyendo esto y en las próximas horas vamos a saber si le sirve a alguien más que a nosotros mismos. "
            "Eso es todo lo que importa hoy.'\n\n"
            "[00:30-00:43] CTA:\n"
            "'Link en bio. Si lo pruebas hoy, cuéntame qué piensas en los comentarios.'\n\n"
            "--- TEXTO PANTALLA ---\n"
            "0:00 → 'Día del lanzamiento'\n"
            "0:15 → 'Meses de trabajo. Horas para saber si valió la pena.'\n"
            "0:35 → 'Link en bio — cuéntame qué te parece'"
        ),
        "direccion_visual": (
            "Video selfie auténtico, mañana del lanzamiento. "
            "Sin producción — solo la cámara frontal y el momento real. "
            "Este es el video más humano de toda la campaña."
        ),
        "hashtags": "#buildinpublic #saas #propfirm #futurestrading #lanzamiento #tradingjournal #fundedtrader",
        "publicacion": "Jueves 08:00 AM hora Colombia/México",
        "hook_alternativo": "Meses de trabajo. Hoy lo lanzamos. Así se siente el día del lanzamiento.",
        "notas_marca": "El video más auténtico de toda la campaña. No editar en exceso. La emoción genuina es el contenido.",
        "estado": "Borrador",
    },
]

# ---------------------------------------------------------------------------
# COLORES
# ---------------------------------------------------------------------------
COLOR_HEADER_BG  = "1A1A2E"   # azul muy oscuro
COLOR_HEADER_FG  = "FFFFFF"   # blanco

COLOR_S1_BG      = "0D1117"   # negro azulado
COLOR_S1_ACCENT  = "00FF87"   # verde neón (semana 1)
COLOR_S2_BG      = "0D1117"
COLOR_S2_ACCENT  = "3B82F6"   # azul (semana 2)
COLOR_S3_BG      = "0D1117"
COLOR_S3_ACCENT  = "A855F7"   # violeta (semana 3)
COLOR_S4_BG      = "0D1117"
COLOR_S4_ACCENT  = "F59E0B"   # ámbar (semana 4)

SEMANA_COLORS = {
    "Semana 1 — Presencia":       ("111827", "00FF87"),
    "Semana 2 — Audiencia activa":("0F172A", "3B82F6"),
    "Semana 3 — Influencers":     ("1A0E2E", "A855F7"),
    "Semana 4 — Pre-launch":      ("1C1007", "F59E0B"),
}

PLATFORM_COLORS = {
    "X / Twitter":  "E7E7E7",
    "Instagram":    "FDEBED",
    "TikTok":       "E8F8F0",
}

ESTADO_COLORS = {
    "Borrador":   "374151",
    "Listo":      "065F46",
    "Publicado":  "1E3A5F",
}

# ---------------------------------------------------------------------------
# BUILD WORKBOOK
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Calendario de Contenido"

# --- Column headers ---
headers = [
    "Semana",
    "Fecha",
    "Plataforma",
    "Pilar",
    "Formato",
    "Hook / Título",
    "Copy completo",
    "Dirección visual",
    "Hashtags",
    "Día / Hora publicación",
    "Notas de marca",
    "Estado",
]

col_widths = [22, 13, 15, 18, 22, 40, 80, 60, 55, 30, 45, 12]

header_fill   = PatternFill("solid", fgColor=COLOR_HEADER_BG)
header_font   = Font(bold=True, color=COLOR_HEADER_FG, size=11, name="Calibri")
header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.row_dimensions[1].height = 32

for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill   = header_fill
    cell.font   = header_font
    cell.alignment = header_align
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# --- Data rows ---
current_semana = None
accent_color = "00FF87"

for row_idx, item in enumerate(content, start=2):
    semana = item["semana"]
    row_bg, acc = SEMANA_COLORS.get(semana, ("111827", "FFFFFF"))

    # Alternate row shade
    if row_idx % 2 == 0:
        row_bg_hex = row_bg
    else:
        # slightly lighter
        row_bg_hex = row_bg

    row_fill = PatternFill("solid", fgColor=row_bg_hex)

    ws.row_dimensions[row_idx].height = 80

    values = [
        item["semana"],
        item["fecha"],
        item["plataforma"],
        item["pilar"],
        item["formato"],
        item["hook_titulo"],
        item["copy"],
        item["direccion_visual"],
        item["hashtags"],
        item["publicacion"],
        item["notas_marca"],
        item["estado"],
    ]

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = row_fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        )

        # Special styling per column
        if col_idx == 1:  # Semana
            cell.font = Font(bold=True, color=acc, size=10, name="Calibri")
        elif col_idx == 3:  # Plataforma
            plat_color = PLATFORM_COLORS.get(item["plataforma"], "CCCCCC")
            cell.fill = PatternFill("solid", fgColor=plat_color)
            cell.font = Font(bold=True, color="111827", size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif col_idx == 6:  # Hook
            cell.font = Font(bold=True, color="F9FAFB", size=10, name="Calibri")
        elif col_idx == 12:  # Estado
            estado_bg = ESTADO_COLORS.get(item["estado"], "374151")
            cell.fill = PatternFill("solid", fgColor=estado_bg)
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.font = Font(color="D1D5DB", size=9, name="Calibri")

# --- Freeze top row ---
ws.freeze_panes = "A2"

# --- Auto-filter ---
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# --- Second sheet: A/B Hooks reference ---
ws2 = wb.create_sheet("Hooks AB Testing")
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 45
ws2.column_dimensions["C"].width = 45
ws2.column_dimensions["D"].width = 15

h2_fill = PatternFill("solid", fgColor="1A1A2E")
h2_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
h2_align = Alignment(horizontal="center", vertical="center")

for col_idx, hdr in enumerate(["Plataforma / Fecha", "Hook A (principal)", "Hook B (alternativo)", "Estado"], start=1):
    cell = ws2.cell(row=1, column=col_idx, value=hdr)
    cell.fill = h2_fill
    cell.font = h2_font
    cell.alignment = h2_align

ws2.row_dimensions[1].height = 28

for row_idx, item in enumerate(content, start=2):
    semana = item["semana"]
    row_bg, acc = SEMANA_COLORS.get(semana, ("111827", "FFFFFF"))
    rfill = PatternFill("solid", fgColor=row_bg)

    ws2.row_dimensions[row_idx].height = 50

    vals = [
        f"{item['plataforma']} | {item['fecha']}",
        item["hook_titulo"],
        item["hook_alternativo"],
        item["estado"],
    ]
    for col_idx, val in enumerate(vals, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = rfill
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if col_idx == 4:
            estado_bg = ESTADO_COLORS.get(item["estado"], "374151")
            cell.fill = PatternFill("solid", fgColor=estado_bg)
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            cell.font = Font(bold=True, color="F9FAFB", size=9, name="Calibri")
        else:
            cell.font = Font(color="9CA3AF", size=9, name="Calibri")

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:D1"

# --- Third sheet: Hashtag clusters ---
ws3 = wb.create_sheet("Clusters de Hashtags")
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 70
ws3.column_dimensions["C"].width = 30

for col_idx, hdr in enumerate(["Categoría", "Hashtags", "Uso recomendado"], start=1):
    cell = ws3.cell(row=1, column=col_idx, value=hdr)
    cell.fill = PatternFill("solid", fgColor="1A1A2E")
    cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")

hashtag_data = [
    ("Prop Firms (nicho)", "#propfirm #FTMO #TopStep #ApexTrader #Tradovate #Uprofit #fundedtrader #fundedaccount", "Siempre incluir 2-3 de estos"),
    ("Futuros / Instrumentos", "#futurestrading #NinjaTrader #Tradovate #MES #ES #NQ #CL #GC", "Posts técnicos o de producto"),
    ("Psicología", "#tradingpsychology #tradingmindset #revengetrading #disciplina", "Posts de psicología y comportamiento"),
    ("Journal / App", "#tradingjournal #tradingapp #journaldetrading #analisisdetrading", "Posts de producto"),
    ("Build in public", "#buildinpublic #indiedev #saas #makersoftwitter", "Posts de BTS y desarrollo"),
    ("LATAM amplio", "#trading #invertir #mercados #finanzas #trader", "Reach general — usar con moderación"),
    ("Marca", "#zentrade #zentradeapp", "Siempre incluir en posts de producto"),
]

for row_idx, (cat, tags, uso) in enumerate(hashtag_data, start=2):
    colors = ["111827", "0F172A", "1A0E2E", "1C1007", "111827", "0F172A", "1A0E2E"]
    rfill = PatternFill("solid", fgColor=colors[row_idx % len(colors)])
    for col_idx, val in enumerate([cat, tags, uso], start=1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = rfill
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.font = Font(color="D1D5DB" if col_idx != 1 else "00FF87", size=9, name="Calibri",
                        bold=(col_idx == 1))
    ws3.row_dimensions[row_idx].height = 45

ws3.freeze_panes = "A2"

# ---------------------------------------------------------------------------
# SAVE
# ---------------------------------------------------------------------------
output_path = r"D:\Development\2 - Zentrade\CONTENT-CALENDAR.xlsx"
wb.save(output_path)
print(f"Archivo generado: {output_path}")
print(f"Total de piezas de contenido: {len(content)}")
